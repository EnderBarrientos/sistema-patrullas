# Módulos estándar de Python
import datetime
import os
from io import BytesIO

# Módulos de terceros (instalados vía pip)
import pandas as pd
from dotenv import load_dotenv
from flask import (Flask, flash, make_response, redirect, render_template,
                    request, send_file, url_for)
from flask_login import (LoginManager, UserMixin, current_user, login_required,
                        login_user, logout_user)
from flask_sqlalchemy import SQLAlchemy
from openpyxl.styles import Font, PatternFill, Alignment # Asegúrate de que Alignment esté importado
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange
from sqlalchemy.exc import IntegrityError 
from sqlalchemy.orm import joinedload
from werkzeug.security import check_password_hash, generate_password_hash
from xhtml2pdf import pisa

load_dotenv()

app = Flask(__name__, static_folder='static')
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY')
# app.config['PDFKIT_CONFIG'] = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

#------------------------------------------------------------------------------#
    # Creación de Modelos
#------------------------------------------------------------------------------#

# Modelo de Usuario
class User(UserMixin, db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    id_number = db.Column(db.String(20), unique=True, nullable=False)
    password = db.Column(db.String(255), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)
    first_login = db.Column(db.Boolean, default=True)
    id_patrol_leader = db.Column(db.Integer, db.ForeignKey('patrol_leaders.id'), nullable=True)

# Modelo de Centro de Votación
class VotingCenter(db.Model):
    __tablename__ = 'voting_centers'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.String(255), nullable=True)
    leaders = db.relationship('PatrolLeader', backref='voting_center', lazy=True)
    members = db.relationship('PatrolMember', backref='member_voting_center', lazy=True)

# Modelo de Identificación
class Identification(db.Model):
    __tablename__ = 'identifications'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.String(255), nullable=True)
    leaders = db.relationship('LeaderIdentification', backref='identification', lazy=True)

# Modelo de Jefe de Patrulla
class PatrolLeader(db.Model):
    __tablename__ = 'patrol_leaders'
    id = db.Column(db.Integer, primary_key=True)
    person_in_charge = db.Column(db.String(100), nullable=False)
    id_number = db.Column(db.String(20), unique=True, nullable=False)
    phone_number = db.Column(db.String(20), nullable=True)
    zone = db.Column(db.String(100), nullable=True)
    other_identification = db.Column(db.String(100), nullable=True)
    id_voting_center = db.Column(db.Integer, db.ForeignKey('voting_centers.id'), nullable=False)
    members = db.relationship('PatrolMember', backref='leader', lazy=True, cascade="all, delete-orphan")
    identifications = db.relationship('LeaderIdentification', backref='leader', lazy=True)
    user = db.relationship('User', backref='leader', uselist=False)

# Modelo de Jefes e Identificaciones
class LeaderIdentification(db.Model):
    __tablename__ = 'leaders_identifications'
    id = db.Column(db.Integer, primary_key=True)
    id_patrol_leader = db.Column(db.Integer, db.ForeignKey('patrol_leaders.id'), nullable=False)
    id_identification = db.Column(db.Integer, db.ForeignKey('identifications.id'), nullable=False)

# Modelo de Patrullero
class PatrolMember(db.Model):
    __tablename__ = 'patrol_members'
    id = db.Column(db.Integer, primary_key=True)
    id_number = db.Column(db.String(20), unique=True, nullable=False)
    name = db.Column(db.String(100), nullable=False)
    phone_number = db.Column(db.String(20), nullable=True)
    direction = db.Column(db.String(200), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    id_patrol_leader = db.Column(db.Integer, db.ForeignKey('patrol_leaders.id'), nullable=False)
    id_voting_center = db.Column(db.Integer, db.ForeignKey('voting_centers.id'), nullable=False)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

def id_number_exist(id_number):
    """Verifica si una cédula ya existe como jefe o patrullero"""
    # Verificar si es jefe
    if PatrolLeader.query.filter_by(id_number=id_number).first():
        return True
    
    # Verificar si es patrullero
    if PatrolMember.query.filter_by(id_number=id_number).first():
        return True
    
    return False

def id_number_valid(id_number, id_member=None):
    """Verifica si una cédula es válida para un nuevo o actualizado patrullero."""
    
    # 1. Buscar si existe un Patrullero con la misma cédula
    member = PatrolMember.query.filter_by(id_number=id_number).first()
    
    if member:
        # Si estamos actualizando el mismo patrullero, es válido
        if id_member and member.id == id_member:
            return True
        # Si es un nuevo patrullero o actualizando a otro, la cédula ya está en uso
        return False
    
    # 2. Verificar si la cédula pertenece a algún Jefe de Patrulla
    if PatrolLeader.query.filter_by(id_number=id_number).first():
        # Si la cédula ya está asignada a un jefe, no es válida para un patrullero
        return False
    
    # 3. Verificar si la cédula pertenece a algún Usuario
    if User.query.filter_by(id_number=id_number).first():
        # Si la cédula ya está asignada a un usuario, no es válida para un patrullero
        return False
        
    # Si la cédula no se encontró en Patrulleros, Jefes, ni Usuarios, entonces es válida
    return True

@app.route('/')
def index():
    return render_template('index.html')

#------------------------------------------------------------------------------#
    # Iniciar Sesión
#------------------------------------------------------------------------------#

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        id_number = request.form['id_number']
        password = request.form['password']
        user = User.query.filter_by(id_number=id_number).first()
        # user = User.query.get(id_number=id_number)
        
        if user and check_password_hash(user.password, password):
            login_user(user)
            flash('Has iniciado sesión correctamente.', 'success')
            
            if user.first_login:
                return redirect(url_for('change_password'))
                
            return redirect(url_for('dashboard'))
        flash('Cédula o contraseña incorrectos.', 'danger')
    return render_template('login.html')

#------------------------------------------------------------------------------#
    # Cambiar Contraseña
#------------------------------------------------------------------------------#

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        actual = request.form['actual']
        new = request.form['new']
        confirmation = request.form['confirmation']
        
        # Validar contraseña actual
        if not check_password_hash(current_user.password, actual):
            flash('Contraseña actual incorrecta.', 'danger')
            return render_template('change_password.html')
        
        # Validar coincidencia de contraseñas
        if new != confirmation:
            flash('Las nuevas contraseñas no coinciden.', 'danger')
            return render_template('change_password.html')
        
        # Validar requisitos de la nueva contraseña
        if len(new) < 8:
            flash('La contraseña debe tener al menos 8 caracteres.', 'danger')
            return render_template('change_password.html')
        
        if not any(c.isupper() for c in new):
            flash('La contraseña debe contener al menos una mayúscula.', 'danger')
            return render_template('change_password.html')
        
        if not any(c.islower() for c in new):
            flash('La contraseña debe contener al menos una minúscula.', 'danger')
            return render_template('change_password.html')
        
        if not any(c.isdigit() for c in new):
            flash('La contraseña debe contener al menos un número.', 'danger')
            return render_template('change_password.html')
        
        # Si pasa todas las validaciones, actualizar la contraseña
        current_user.password = generate_password_hash(new, method='pbkdf2:sha256')
        current_user.first_login = False
        db.session.commit()
        
        flash('Contraseña actualizada exitosamente.', 'success')
        return redirect(url_for('dashboard'))
    
    mensaje = "Por seguridad, debe cambiar su contraseña inicial." if current_user.first_login else ""
    return render_template('change_password.html', primer_inicio=current_user.first_login, mensaje=mensaje)

#------------------------------------------------------------------------------#
    # Cerrar Sesión
#------------------------------------------------------------------------------#

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Has cerrado sesión correctamente.', 'success')
    return redirect(url_for('index'))

#------------------------------------------------------------------------------#
    # Panel de Control
#------------------------------------------------------------------------------#

@app.route('/dashboard')
@login_required
def dashboard():
    if current_user.first_login:
        return redirect(url_for('change_password'))
    
    if current_user.leader:
        return render_template('dashboard.html', leader=current_user.leader)
    
    return render_template('register_pending.html')

#------------------------------------------------------------------------------#
    # Administración
#------------------------------------------------------------------------------#

@app.route('/admin', methods=['GET', 'POST'])
@login_required
def admin_site():
    if current_user.first_login:
        return redirect(url_for('change_password'))
    
    if not current_user.is_admin:
        flash('Acceso no autorizado', 'danger')
        return redirect(url_for('dashboard'))
    
    return render_template('admin_site.html')

#------------------------------------------------------------------------------#
    # Administrar Usuarios
#------------------------------------------------------------------------------#

@app.route('/admin/user', methods=['GET', 'POST'])
@login_required
def admin_user():
    
    if not current_user.is_admin:
        flash('Acceso no autorizado', 'danger')
        return redirect(url_for('dashboard'))
    
    # Filtrar usuarios excluyendo al usuario actual
    users = User.query.filter(User.id != current_user.id).all()
    
    return render_template('admin_user.html', users=users)

@app.route('/admin/user/create', methods=['GET', 'POST'])
@login_required
def create_user():
    if not current_user.is_admin:
        flash('Acceso no autorizado', 'danger')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        id_number = request.form['id_number']
        name = request.form['name']
        is_admin = 'is_admin' in request.form
        
        # Verificar si la cédula ya existe
        if User.query.filter_by(id_number=id_number).first():
            flash('El usuario con esta cédula ya existe.', 'danger')
            return redirect(url_for('create_user'))
        
        hashed_pw = generate_password_hash('123456', method='pbkdf2:sha256')
        new_user = User(
            id_number=id_number,
            password=hashed_pw,
            name=name,
            is_admin=is_admin,
            first_login=True
        )
        db.session.add(new_user)
        db.session.commit()
        
        flash(f'Usuario {name} creado exitosamente. Contraseña inicial: 123456', 'success')
        return redirect(url_for('admin_user'))
    
    return render_template('create_user.html')


#------------------------------------------------------------------------------#
    # Administrar Jefes de Patrullas
#------------------------------------------------------------------------------#

@app.route('/admin/patrol-leader', methods=['GET', 'POST'])
@login_required
def admin_patrol_leader():
    
    if not current_user.is_admin:
        flash('Acceso no autorizado', 'danger')
        return redirect(url_for('dashboard'))
    
    patrol_leaders = PatrolLeader.query.join(VotingCenter).add_entity(VotingCenter).all()
    
    return render_template('admin_patrol_leader.html', leaders=patrol_leaders)

#------------------------------------------------------------------------------#
    # Administrar Centros de Votación
#------------------------------------------------------------------------------#

@app.route('/admin/voting-center', methods=['GET', 'POST'])
@login_required
def admin_voting_center():
    
    if not current_user.is_admin:
        flash('Acceso no autorizado', 'danger')
        return redirect(url_for('dashboard'))
    
    voting_centers = VotingCenter.query.all()
    
    return render_template('admin_voting_center.html', centers=voting_centers)

@app.route('/admin/voting-center/create', methods=['GET', 'POST'])
@login_required
def create_voting_center():
    if not current_user.is_admin:
        flash('Acceso no autorizado', 'danger')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        name = request.form['name']
        description = request.form['description']
        
        # Verificar si la cédula ya existe
        if VotingCenter.query.filter_by(name=name.lower()).first():
            flash('Este centro de votación ya existe.', 'danger')
            return redirect(url_for('create_voting_center'))
        
        new_center = VotingCenter(
            name=name.lower(), 
            description=description
        )
        db.session.add(new_center)
        db.session.commit()
        
        flash(f'Centro de votación {name} creado exitosamente.', 'success')
        return redirect(url_for('admin_voting_center'))
    
    return render_template('create_voting_center.html')

@app.route('/delete-center/<int:id_center>', methods=['POST'])
@login_required
def delete_center(id_center):
    center = VotingCenter.query.get_or_404(id_center)
    
    # Verificar permisos
    if not current_user.is_admin:
        flash('No tienes permiso para eliminar este centro', 'danger')
        return redirect(url_for('dashboard'))
    
    # Verificar asociaciones
    has_leaders = PatrolLeader.query.filter_by(id_voting_center=id_center).first() is not None
    has_members = PatrolMember.query.filter_by(id_voting_center=id_center).first() is not None
    
    if has_leaders or has_members:
        leader_count = PatrolLeader.query.filter_by(id_voting_center=id_center).count()
        member_count = PatrolMember.query.filter_by(id_voting_center=id_center).count()
        
        flash_message = (
            f'No se puede eliminar el centro "{center.name}" porque está asociado a: '
            f'{leader_count} jefe(s) de patrulla y {member_count} patrullero(s). '
            'Primero debe reasignar o eliminar estos registros.'
        )
        flash(flash_message, 'danger')
        return redirect(url_for('admin_voting_center'))
    
    try:
        db.session.delete(center)
        db.session.commit()
        flash('Centro eliminado exitosamente.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar centro de votación: {str(e)}', 'danger')
    
    return redirect(url_for('admin_voting_center'))
#------------------------------------------------------------------------------#
    # Administrar Identificaciones
#------------------------------------------------------------------------------#

@app.route('/admin/identification', methods=['GET', 'POST'])
@login_required
def admin_identification():
    
    if not current_user.is_admin:
        flash('Acceso no autorizado', 'danger')
        return redirect(url_for('dashboard'))
    
    identifications = Identification.query.all()
    
    return render_template('admin_identification.html', identifications=identifications)

@app.route('/admin/identification/create', methods=['GET', 'POST'])
@login_required
def create_identification():
    if not current_user.is_admin:
        flash('Acceso no autorizado', 'danger')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        name = request.form['name']
        description = request.form['description']
        
        if Identification.query.filter_by(name=name.lower()).first():
            flash('Esta identificación ya existe.', 'danger')
            return redirect(url_for('create_identification'))
        
        new_identification = Identification(
            name=name.lower(), 
            description=description
        )
        db.session.add(new_identification)
        db.session.commit()
        
        flash(f'Identificación {name} creado exitosamente.', 'success')
        return redirect(url_for('admin_identification'))
    
    return render_template('create_identification.html')


@app.route('/delete-identification/<int:id_identification>', methods=['POST'])
@login_required
def delete_identification(id_identification):
    identification = Identification.query.get_or_404(id_identification)
    
    # Verificar permisos
    if not current_user.is_admin:
        flash('No tienes permiso para eliminar esta identificación', 'danger')
        return redirect(url_for('dashboard'))
    
    # Verificar asociaciones con jefes de patrulla
    has_leader_associations = LeaderIdentification.query.filter_by(id_identification=id_identification).first() is not None
    
    if has_leader_associations:
        association_count = LeaderIdentification.query.filter_by(id_identification=id_identification).count()
        
        flash_message = (
            f'No se puede eliminar la identificación "{identification.name}" porque está asociada a: '
            f'{association_count} jefe(s) de patrulla. '
            'Primero debe reasignar o eliminar estas asociaciones.'
        )
        flash(flash_message, 'danger')
        return redirect(url_for('admin_identification'))  # Asegúrate de tener esta ruta definida
    
    try:
        # Eliminar la identificación
        db.session.delete(identification)
        db.session.commit()
        flash('Identificación eliminada exitosamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar la identificación: {str(e)}', 'danger')
    
    return redirect(url_for('admin_identification'))  # Redirigir a la vista de administración de identificaciones

#------------------------------------------------------------------------------#
    # Registrar Jefe de Patrulla
#------------------------------------------------------------------------------#

@app.route('/register-leader', methods=['GET', 'POST'])
@login_required
def register_leader():
    if current_user.first_login:
        return redirect(url_for('change_password'))
    
    if current_user.leader:
        flash('Ya eres jefe de una patrulla.', 'warning')
        return redirect(url_for('dashboard'))
    
    # Obtener todos los centros de votación
    voting_centers = VotingCenter.query.order_by(VotingCenter.name).all()
    
    identifications = Identification.query.order_by(Identification.name).all()
    
    if request.method == 'POST':
        try:
            id_number_leader = request.form['id_number_leader']
            
            # Verificar si la cédula del jefe ya existe
            if id_number_exist(id_number_leader):
                flash('La cédula del jefe ya está registrada como patrullero o en otra patrulla', 'danger')
                return redirect(url_for('register_leader'))
            
            # Registrar jefe de patrulla
            leader = PatrolLeader(
                person_in_charge=request.form['person_in_charge'],
                id_number=id_number_leader,
                phone_number="N/A" if not request.form.get('phone_number') else request.form['phone_number'],
                zone="N/A" if not request.form.get('zone') else request.form['zone'],
                id_voting_center=request.form['id_voting_center'],
                other_identification = request.form['other_identification'],
            )
            db.session.add(leader)
            db.session.flush()  # Obtener ID sin commit
            
            # Guardar las identificaciones seleccionadas
            selected_identifications = request.form.getlist('identifications[]')
            for identification_id in selected_identifications:
                leader_identification = LeaderIdentification(
                    id_patrol_leader=leader.id,
                    id_identification=identification_id
                )
                db.session.add(leader_identification)
            
            # --- Registrar patrulleros ---
            i = 0
            cedulas_registradas_en_formulario = set() # Para detectar duplicados dentro del mismo formulario

            while True:
                member_id_number = request.form.get(f'member_id_number_{i}', '').strip()
                member_name = request.form.get(f'member_name_{i}', '').strip()
                member_phone = request.form.get(f'member_phone_{i}', '').strip()
                member_direction = request.form.get(f'member_direction_{i}', '').strip()
                member_voting_center_id = request.form.get(f'member_id_voting_center_{i}', '').strip()

                # Si no hay datos para esta fila, asumimos que no hay más patrulleros
                if not member_id_number and not member_name:
                    break

                # Validaciones
                if not member_id_number:
                    flash(f'La cédula del patrullero en la fila {i+1} es obligatoria.', 'danger')
                    db.session.rollback()
                    return redirect(url_for('register_leader'))

                if not member_name:
                    flash(f'El nombre del patrullero en la fila {i+1} es obligatorio.', 'danger')
                    db.session.rollback()
                    return redirect(url_for('register_leader'))

                # Verificar si la cédula está duplicada en el mismo formulario
                if member_id_number in cedulas_registradas_en_formulario:
                    flash(f'La cédula {member_id_number} está duplicada en el formulario (fila {i+1}).', 'danger')
                    db.session.rollback()
                    return redirect(url_for('register_leader'))
                cedulas_registradas_en_formulario.add(member_id_number)

                # Verificar si la cédula del patrullero ya existe en la base de datos (como líder o como otro patrullero)
                if id_number_exist(member_id_number):
                    flash(f'La cédula {member_id_number} ya está registrada como patrullero o jefe (fila {i+1}).', 'danger')
                    db.session.rollback()
                    return redirect(url_for('register_leader'))

                # Asignar centro de votación si no se proporciona uno específico para el miembro
                final_member_voting_center_id = member_voting_center_id if member_voting_center_id else leader.id_voting_center

                patrol_member = PatrolMember(
                    id_patrol_leader=leader.id,
                    id_number=member_id_number,
                    name=member_name,
                    phone_number=member_phone if member_phone else "N/A",
                    direction=member_direction if member_direction else "N/A",
                    id_voting_center=final_member_voting_center_id
                )
                db.session.add(patrol_member)
                i += 1
            # --- Fin del registro de patrulleros ---
            
            # Asociar el jefe al usuario actual
            current_user.id_patrol_leader = leader.id
            db.session.commit()
            
            flash('¡Registro de patrulla completado exitosamente!', 'success')
            return redirect(url_for('show_patrol', id_leader=leader.id))
        
        except IntegrityError:
            db.session.rollback()
            flash('Error: La cédula del jefe ya está registrada', 'danger')
            return redirect(url_for('register_leader'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error inesperado: {str(e)}', 'danger')
            return redirect(url_for('register_leader'))
    
    return render_template('register_leader.html', user=current_user, voting_centers=voting_centers, identifications=identifications)

#------------------------------------------------------------------------------#
    # Mostrar Patrulla
#------------------------------------------------------------------------------#

@app.route('/show-patrol/<int:id_leader>')
@login_required
def show_patrol(id_leader):
    # Verificar que el usuario tenga acceso a esta patrulla
    if not current_user.is_admin and (not current_user.leader or current_user.leader.id != id_leader):
        flash('No tienes permiso para ver esta patrulla', 'danger')
        return redirect(url_for('dashboard'))
    identifications = LeaderIdentification.query.filter_by(id_patrol_leader=id_leader).all()
    
    leader = PatrolLeader.query.get_or_404(id_leader)
    
    members = PatrolMember.query.filter_by(id_patrol_leader=id_leader).order_by(PatrolMember.created_at).all()
    
    return render_template('show_patrol.html', leader=leader, identifications=identifications, members=members)

#------------------------------------------------------------------------------#
    # Eliminar Patrulla
#------------------------------------------------------------------------------#

@app.route('/delete-leader/<int:id_leader>', methods=['POST'])
@login_required
def delete_leader(id_leader):
    # if not current_user.is_admin:
    #     flash('No tienes permisos para realizar esta acción', 'danger')
    #     return redirect(url_for('dashboard'))
    
    leader = PatrolLeader.query.get_or_404(id_leader)
    
    try:
        # Desvincular al usuario asociado
        if leader.user:
            leader.user.id_patrol_leader = None
            
        # Eliminar las identificaciones del líder
        LeaderIdentification.query.filter_by(id_patrol_leader=leader.id).delete()
        
        # Finalmente eliminar el líder
        db.session.delete(leader)
        
        db.session.commit()
        flash('Patrulla eliminada exitosamente.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar la patrulla: {str(e)}', 'danger')
        app.logger.error(f'Error deleting leader {id_leader}: {str(e)}')
    
    return redirect(url_for('dashboard'))

#------------------------------------------------------------------------------#
    # Agregar Patrulleros
#------------------------------------------------------------------------------#

@app.route('/add_member/<int:id_leader>', methods=['GET', 'POST'])
@login_required
def add_member(id_leader):
    # Verificar permisos
    if not current_user.is_admin and (not current_user.leader or current_user.leader.id != id_leader):
        flash('No tienes permiso para agregar patrulleros a esta patrulla', 'danger')
        return redirect(url_for('dashboard'))
    
    leader = PatrolLeader.query.get_or_404(id_leader)
    
    # Verificar que no se exceda el límite de 21 patrulleros
    if len(leader.members) >= 21:
        flash('No se pueden agregar más de 21 patrulleros', 'danger')
        return redirect(url_for('show_patrol', id_leader= id_leader))
    
    # Obtener todos los centros de votación
    voting_centers = VotingCenter.query.order_by(VotingCenter.name).all()
    
    if request.method == 'POST':
        try:
            id_number = request.form['id_number'].strip()
            name = request.form['name'].strip()
            
            phone_number = request.form.get('phone_number', '').strip()
            if not phone_number: # Si después de stripear sigue vacío
                phone_number = 'N/A'
            
            direction = request.form.get('direction', '').strip()
            if not direction: # Si después de stripear sigue vacío
                direction = 'N/A'
            
            # Verificar si la cédula es válida
            if not id_number_valid(id_number):
                flash(f'La cédula {id_number} ya está registrada en el sistema', 'danger')
                return render_template('form_member.html', leader=leader, voting_centers=voting_centers)
            
            # Crear nuevo patrullero
            member = PatrolMember(
                id_number=id_number,
                name=name,
                phone_number=phone_number,
                direction=direction,
                id_patrol_leader=leader.id,
                id_voting_center=request.form.get('id_voting_center')
            )
            db.session.add(member)
            db.session.commit()
            
            flash('Patrullero agregado exitosamente', 'success')
            return redirect(url_for('show_patrol', id_leader=id_leader))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al agregar patrullero: {str(e)}', 'danger')
    
    return render_template('form_member.html', leader=leader, voting_centers=voting_centers)

@app.route('/edit-member/<int:id_member>', methods=['GET', 'POST'])
@login_required
def edit_member(id_member):
    member = PatrolMember.query.get_or_404(id_member)
    id_leader = member.id_patrol_leader
    
    # Obtener todos los centros de votación
    voting_centers = VotingCenter.query.order_by(VotingCenter.name).all()
    
    # Verificar permisos
    if not current_user.is_admin and (not current_user.leader or current_user.leader.id != id_leader):
        flash('No tienes permiso para eliminar este patrullero', 'danger')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        try:
            id_number = request.form['id_number'].strip()
            name = request.form['name'].strip()
            
            # Verificar si la cédula es válida
            if not id_number_valid(id_number, member.id):
                flash(f'La cédula {id_number} ya está registrada en el sistema', 'danger')
                return render_template('form_member.html', leader=member.leader, member=member, voting_centers=voting_centers)
            
            # Actualizar datos
            member.id_number = id_number
            member.name = name
            member.phone_number = request.form.get('phone_number')
            member.direction = request.form.get('direction')
            member.id_voting_center = request.form.get('id_voting_center')
            
            db.session.commit()
            flash('Patrullero actualizado exitosamente', 'success')
            return redirect(url_for('show_patrol', id_leader=id_leader))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar patrullero: {str(e)}', 'danger')
    
    return render_template('form_member.html', leader=member.leader, member=member, voting_centers=voting_centers)

@app.route('/delete-member/<int:id_member>', methods=['POST'])
@login_required
def delete_member(id_member):
    member = PatrolMember.query.get_or_404(id_member)
    id_leader = member.id_patrol_leader
    
    # Verificar permisos
    if not current_user.is_admin and (not current_user.leader or current_user.leader.id != id_leader):
        flash('No tienes permiso para eliminar este patrullero', 'danger')
        return redirect(url_for('dashboard'))
    
    try:
        db.session.delete(member)
        db.session.commit()
        flash('Patrullero eliminado exitosamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar patrullero: {str(e)}', 'danger')
    
    return redirect(url_for('show_patrol', id_leader=id_leader))

@app.route('/export-pdf/<int:id_leader>')
@login_required
def export_pdf(id_leader):
    leader = PatrolLeader.query.get_or_404(id_leader)
    members = PatrolMember.query.filter_by(id_patrol_leader=id_leader).order_by(PatrolMember.created_at).all()
    identifications = LeaderIdentification.query.filter_by(id_patrol_leader=id_leader).all()

    if not leader:
        flash(f'Jefe de Patrulla con ID {id_leader} no encontrado.', 'danger')
        return redirect(url_for('dashboard'))

    # Renderizar plantilla HTML
    rendered = render_template('print_patrol.html', leader=leader, members=members, identifications=identifications)

    # Crear el archivo PDF en memoria
    pdf_buffer = BytesIO()
    try:
        # Pasar el HTML renderizado a pisa.CreatePDF
        pisa_status = pisa.CreatePDF(
            rendered,                # el contenido HTML a convertir
            dest=pdf_buffer)         # el buffer de bytes para guardar el PDF

        # Si hubo errores en la generación del PDF
        if pisa_status.err:
            flash(f'Error al generar el PDF: {pisa_status.err}', 'danger')
            return redirect(url_for('show_patrol', id_leader=id_leader))

    except Exception as e:
        flash(f'Error inesperado al generar el PDF: {str(e)}', 'danger')
        return redirect(url_for('show_patrol', id_leader=id_leader))

    # Mover el puntero al inicio del buffer para leer su contenido
    pdf_buffer.seek(0)
    pdf = pdf_buffer.getvalue()

    # Crear respuesta con el PDF
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    filename = f"patrulla_{leader.person_in_charge}_{id_leader}.pdf".replace(" ", "_")
    response.headers['Content-Disposition'] = f'inline; filename={filename}'
    return response

@app.route('/print-patrol/<int:id_leader>')
@login_required
def print_patrol(id_leader):
    if not current_user.is_admin and (not current_user.leader or current_user.leader.id != id_leader):
        flash('No tienes permiso para imprimir esta patrulla', 'danger')
        return redirect(url_for('dashboard'))
    
    return redirect(url_for('export_pdf', id_leader=id_leader))

@app.route('/admin/export-excel')
@login_required
def export_excel():
    if not current_user.is_admin:
        flash('Acceso no autorizado', 'danger')
        return redirect(url_for('dashboard'))

    try:
        # Obtener datos de jefes ORDENADOS POR ID
        leaders = PatrolLeader.query.options(
            joinedload(PatrolLeader.voting_center),
            joinedload(PatrolLeader.user),
            joinedload(PatrolLeader.identifications).joinedload(LeaderIdentification.identification)
        ).order_by(PatrolLeader.id).all()

        # Procesar datos de jefes
        leaders_data = []
        for leader in leaders:
            identification_names = []
            for li in leader.identifications:
                if li.identification:
                    identification_names.append(li.identification.name)
            if leader.other_identification:
                identification_names.append(f"OTRO: {leader.other_identification}")

            leaders_data.append({
                'ID Jefe': leader.id,
                'Responsable': leader.person_in_charge,
                'Cédula Jefe': leader.id_number,
                'Teléfono Jefe': leader.phone_number,
                'Zona/Sector Jefe': leader.zone,
                'Centro de Votación Jefe': leader.voting_center.name if leader.voting_center else 'N/A',
                'Identificaciones': ", ".join(identification_names) if identification_names else 'Ninguna',
                'Usuario Asociado': leader.user.name if leader.user else 'Sin usuario'
            })

        # Obtener datos de patrulleros ORDENADOS POR ID
        members = PatrolMember.query.options(
            joinedload(PatrolMember.member_voting_center),
            joinedload(PatrolMember.leader)
        ).order_by(PatrolMember.id).all()

        # Procesar datos de patrulleros
        members_data = []
        for member in members:
            members_data.append({
                'ID Patrullero': member.id,
                'Cédula Patrullero': member.id_number,
                'Nombres y Apellidos Patrullero': member.name,
                'Teléfono Patrullero': member.phone_number,
                'Dirección Patrullero': member.direction,
                'Centro de Votación Patrullero': member.member_voting_center.name if member.member_voting_center else 'N/A',
                'Fecha Creación Patrullero': member.created_at.strftime("%Y-%m-%d %H:%M:%S") if member.created_at else 'N/A',
                'ID Jefe Asociado': member.id_patrol_leader,
                'Responsable Jefe Asociado': member.leader.person_in_charge if member.leader else 'N/A',
                'Cédula Jefe Asociado': member.leader.id_number if member.leader else 'N/A'
            })

        # Obtener datos de usuarios ORDENADOS POR ID
        users = User.query.options(joinedload(User.leader)).order_by(User.id).all()

        # Procesar datos de usuarios
        users_data = [{
            'ID Usuario': user.id,
            'Cédula Usuario': user.id_number,
            'Nombre Usuario': user.name,
            'Es Administrador': 'Sí' if user.is_admin else 'No',
            'Primera Sesión': 'Sí' if user.first_login else 'No',
            'ID Jefe Asociado (Usuario)': user.id_patrol_leader,
            'Responsable Jefe Asociado (Usuario)': user.leader.person_in_charge if user.leader else 'Sin patrulla'
        } for user in users]

        # Crear DataFrames (ya estarán ordenados por la consulta SQL)
        leaders_df = pd.DataFrame(leaders_data)
        members_df = pd.DataFrame(members_data)
        users_df = pd.DataFrame(users_data)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            sheets_info = { # Renamed to sheets_info for clarity
                'Jefes de Patrulla': leaders_df,
                'Patrulleros': members_df,
                'Usuarios': users_df
            }

            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            title_font = Font(bold=True, size=16)
            title_alignment = Alignment(horizontal="center", vertical="center") # Alignment for title

            # Define alignment for data cells
            data_alignment = Alignment(horizontal="center", vertical="center") # New alignment for data

            first_sheet_written = True

            for sheet_name, df in sheets_info.items():
                if not df.empty:
                    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

                    workbook = writer.book
                    worksheet = writer.sheets[sheet_name]

                    # --- Add Title ---
                    last_col_letter = get_column_letter(len(df.columns))
                    worksheet.merge_cells(f'A1:{last_col_letter}1')
                    title_cell = worksheet['A1']
                    title_cell.value = f"REPORTE DE {sheet_name.upper()}"
                    title_cell.font = title_font
                    title_cell.alignment = title_alignment
                    worksheet.row_dimensions[1].height = 30

                    # Apply header style and adjust column widths
                    for col_idx, column in enumerate(df.columns, 1):
                        header_cell = worksheet.cell(row=2, column=col_idx) # Headers are in row 2
                        header_cell.font = header_font
                        header_cell.fill = header_fill
                        header_cell.alignment = data_alignment # Also center the header text

                        max_length = max(
                            len(str(df.columns[col_idx-1])), # Header text length
                            *[len(str(x)) for x in df.iloc[:, col_idx-1].astype(str)] # Max length of data
                        )
                        worksheet.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

                    # --- Center all data cells (from row 3 onwards) ---
                    # We start from row 3 because row 1 is the title and row 2 is the header
                    for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row,
                                                min_col=1, max_col=worksheet.max_column):
                        for cell in row:
                            cell.alignment = data_alignment

                    worksheet.freeze_panes = 'A3'
                    worksheet.sheet_state = 'visible'

                    if first_sheet_written:
                        workbook.active = workbook.index(worksheet)
                        first_sheet_written = False
                else:
                    # If this DataFrame is empty, but others might exist or be created
                    pass

            # Critical Check: If no sheets were added because all DataFrames were empty
            if not workbook.sheetnames:
                ws = workbook.create_sheet("Sin Datos", 0)
                ws['A1'] = "No hay datos disponibles para exportar."
                ws['A1'].font = Font(bold=True, size=14)
                ws.merge_cells('A1:C1')
                ws.sheet_state = 'visible'
                workbook.active = 0

        output.seek(0)

        fecha = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reporte_patrullas_{fecha}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        # app.logger.error(f"Error generating Excel report: {e}", exc_info=True)
        flash(f'Error al generar el reporte Excel: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))
    
# Error 404 - Página no encontrada
@app.errorhandler(404)
def page_not_found(error):
    return render_template(
        'error_base.html',
        error_code=404,
        error_name="Página no encontrada",
        error_description="La URL que buscas no existe. ¿Quizás un error de escritura?",
        icon="🔍"
    ), 404

# Error 405 - Método no permitido
@app.errorhandler(405)
def method_not_allowed(error):
    return render_template(
        'error_base.html',
        error_code=405,
        error_name="Método no permitido",
        error_description="Oops! Esta acción no está permitida.",
        icon="🚫"
    ), 405

# Error 500 - Error del servidor
@app.errorhandler(500)
def internal_error(error):
    return render_template(
        'error_base.html',
        error_code=500,
        error_name="Error del servidor",
        error_description="Algo salió mal. Por favor, intenta más tarde.",
        icon="💥"
    ), 500

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        
        # Crear usuario admin si no existe
        if not User.query.filter_by(id_number='admin').first():
            hashed_pw = generate_password_hash('123456', method='pbkdf2:sha256')
            admin = User(
                id_number='admin', 
                password=hashed_pw, 
                name='Administrador',
                is_admin=True, 
                first_login=True
            )
            db.session.add(admin)
            db.session.commit()
            print('Usuario admin creado: cedula=admin, password=123456')
    
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
